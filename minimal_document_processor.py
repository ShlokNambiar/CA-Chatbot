"""
Minimal Document Processor - Lightweight version without heavy dependencies
"""
import os
from typing import List, Dict, Any
import PyPDF2
from docx import Document
import openpyxl
import io

class MinimalDocumentProcessor:
    """Lightweight document processor without pandas or pdfplumber"""
    
    def __init__(self):
        self.supported_formats = ['.pdf', '.docx', '.xlsx', '.txt']
    
    def process_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF using PyPDF2"""
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            text_content = []
            for page in pdf_reader.pages:
                text_content.append(page.extract_text())
            
            return "\n".join(text_content)
            
        except Exception as e:
            print(f"Error processing PDF: {e}")
            return ""
    
    def process_docx(self, file_content: bytes) -> str:
        """Extract text from Word document"""
        try:
            doc_file = io.BytesIO(file_content)
            doc = Document(doc_file)
            
            text_content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)
            
            return "\n".join(text_content)
            
        except Exception as e:
            print(f"Error processing DOCX: {e}")
            return ""
    
    def process_xlsx(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            text_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        text_content.append(" | ".join(row_text))
            
            return "\n".join(text_content)
            
        except Exception as e:
            print(f"Error processing XLSX: {e}")
            return ""
    
    def process_txt(self, file_content: bytes) -> str:
        """Process plain text file"""
        try:
            return file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_content.decode('latin-1')
            except Exception as e:
                print(f"Error processing TXT: {e}")
                return ""
    
    def process_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process a file and extract text content"""
        file_ext = os.path.splitext(filename.lower())[1]
        
        if file_ext not in self.supported_formats:
            return {
                "success": False,
                "error": f"Unsupported file format: {file_ext}",
                "text": ""
            }
        
        try:
            if file_ext == '.pdf':
                text = self.process_pdf(file_content)
            elif file_ext == '.docx':
                text = self.process_docx(file_content)
            elif file_ext == '.xlsx':
                text = self.process_xlsx(file_content)
            elif file_ext == '.txt':
                text = self.process_txt(file_content)
            else:
                return {
                    "success": False,
                    "error": f"Handler not implemented for {file_ext}",
                    "text": ""
                }
            
            return {
                "success": True,
                "text": text,
                "filename": filename,
                "format": file_ext,
                "length": len(text)
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "text": ""
            }
    
    def chunk_text(self, text: str, chunk_size: int = 1000, overlap: int = 100) -> List[str]:
        """Split text into chunks for processing"""
        if len(text) <= chunk_size:
            return [text]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + chunk_size
            
            # Try to break at sentence or paragraph boundaries
            if end < len(text):
                # Look for sentence endings
                for i in range(end, max(start + chunk_size - 200, start), -1):
                    if text[i] in '.!?\n':
                        end = i + 1
                        break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - overlap
            if start >= len(text):
                break
        
        return chunks
